from flask import Flask, jsonify, request, send_file
import io
import logging
import torch
from diffusers import StableDiffusionPipeline
import warnings
import openpyxl
from openpyxl.drawing.image import Image
import os
import tempfile

# Initialize Flask app
app = Flask(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Suppress unnecessary warnings
warnings.filterwarnings("ignore", category=UserWarning)

model_id = "runwayml/stable-diffusion-v1-5"
pipe = None
device = "cpu"

def init_model():
    """Initialize the Stable Diffusion model lazily."""
    global pipe
    if pipe is None:
        try:
            logger.info(f"Loading model {model_id} on {device}")
            pipe = StableDiffusionPipeline.from_pretrained(model_id)
            pipe = pipe.to(device)
            logger.info("Model loaded successfully")
        except Exception as e:
            logger.error(f"Failed to load model: {str(e)}")
            raise RuntimeError(f"Model initialization failed: {str(e)}")

def generate_images(prompt: str, num_images: int = 5):
    """Generate multiple images from a text prompt."""
    try:
        init_model()
        images = []

        # Generate multiple images
        for i in range(num_images):
            logger.info(f"Generating image {i+1} for prompt: {prompt}")
            result = pipe(prompt, num_inference_steps=50, guidance_scale=7.5)
            image = result.images[0]

            if hasattr(result, 'nsfw_content_detected') and any(result.nsfw_content_detected):
                logger.warning(f"NSFW content detected in image {i+1}")
                continue  # Skip NSFW images

            images.append(image)

        if not images:
            raise ValueError("All generated images contained NSFW content")

        return images
    except Exception as e:
        logger.error(f"Image generation failed: {str(e)}")
        raise RuntimeError(f"Image generation failed: {str(e)}")

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data = request.json
        if not data or 'prompt' not in data:
            return jsonify({'error': 'Prompt is required'}), 400

        prompt = data.get('prompt')
        if not isinstance(prompt, str) or not prompt.strip():
            return jsonify({'error': 'Prompt must be a non-empty string'}), 400

        # Get the number of images from the request body
        num_images = int(data.get('numberOfImages', 5))
        logger.info(f'Generating {num_images} image(s) in total...')
        # Generate five images
        images = generate_images(prompt, num_images=num_images)

        # Create an Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Generated Images"

        # Add the prompt to the first cell
        ws['A1'] = "Prompt"
        ws['B1'] = prompt

        # Save images to temporary files and embed them in the Excel sheet
        temp_image_paths = []
        for idx, image in enumerate(images):
            # Save each image to a temporary file
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_img_file:
                image.save(temp_img_file, format='PNG')
                temp_image_paths.append(temp_img_file.name)

            # Embed the image in the Excel sheet
            img = Image(temp_image_paths[-1])
            # Place images vertically, starting at A2, with some spacing
            cell = f'A{2 + idx * 20}'  # Adjust row offset for spacing (assuming ~20 rows per image)
            ws.add_image(img, cell)

        # Save the Excel file to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel_file:
            wb.save(temp_excel_file.name)
            temp_excel_path = temp_excel_file.name

        # Clean up temporary image files
        for temp_img_path in temp_image_paths:
            try:
                os.remove(temp_img_path)
            except Exception as e:
                logger.error(f"Failed to delete temporary image file {temp_img_path}: {str(e)}")

        # Send the Excel file as a response
        return send_file(
            temp_excel_path,
            as_attachment=True,
            download_name='generated_images.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ValueError as ve:
        logger.warning(f"Generation error: {str(ve)}")
        return jsonify({'error': str(ve)}), 400
    except Exception as e:
        logger.error(f"Error in generate endpoint: {str(e)}")
        return jsonify({'error': 'Failed to generate images'}), 500
    finally:
        # Clean up the temporary Excel file if it exists
        if 'temp_excel_path' in locals() and os.path.exists(temp_excel_path):
            try:
                os.remove(temp_excel_path)
            except Exception as e:
                logger.error(f"Failed to delete temporary Excel file: {str(e)}")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)